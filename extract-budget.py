import openpyxl

wb = openpyxl.load_workbook('warren-latest.xlsx', data_only=True)

# Check YTD Budget Totals sheet
print("=== YTD Budget Totals ===")
ws = wb['YTD Budget Totals']
print(f"Max rows: {ws.max_row}, Max cols: {ws.max_column}")
for row in range(1, min(30, ws.max_row + 1)):
    vals = []
    for col in range(1, min(20, ws.max_column + 1)):
        v = ws.cell(row, col).value
        if v is not None:
            vals.append(f"C{col}={v}")
    if vals:
        print(f"Row {row}: {', '.join(vals[:12])}")

# Check if there's a per-period budget breakdown
print("\n\nSearching for 'budget' and 'net' in all sheets...")
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    for row in range(1, min(250, ws.max_row + 1)):
        label = ws.cell(row, 1).value
        if label and 'budget' in str(label).lower() and 'net' in str(label).lower():
            vals = [ws.cell(row, c).value for c in range(1, 16)]
            print(f"  [{sheet_name}] Row {row}: {vals}")

# Also look at the Income sheet for budget
print("\n\n=== Income Sheet (first 10 rows) ===")
ws = wb['Income']
for row in range(1, min(15, ws.max_row + 1)):
    vals = []
    for col in range(1, min(20, ws.max_column + 1)):
        v = ws.cell(row, col).value
        if v is not None:
            vals.append(f"C{col}={v}")
    if vals:
        print(f"Row {row}: {', '.join(vals[:12])}")

# Check the Variance Analysis sheet for what Warren sees
print("\n\n=== Variance Analysis Sheet ===")
ws = wb['Variance Analysis']
print(f"Max rows: {ws.max_row}")
for row in range(1, min(30, ws.max_row + 1)):
    vals = []
    for col in range(1, min(20, ws.max_column + 1)):
        v = ws.cell(row, col).value
        if v is not None:
            vals.append(f"C{col}={v}")
    if vals:
        print(f"Row {row}: {', '.join(vals[:12])}")

# Find Net Income in Variance Analysis
print("\nSearching for 'net income' in Variance Analysis...")
for row in range(1, ws.max_row + 1):
    label = ws.cell(row, 1).value
    if label and 'net income' in str(label).lower():
        vals = [ws.cell(row, c).value for c in range(1, 20)]
        print(f"Row {row}: {vals}")
